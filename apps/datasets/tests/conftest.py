from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl.workbook import Workbook


@pytest.fixture
def excel_file():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "name"
    ws["B1"] = "age"
    ws["A2"] = "Egor"
    ws["B2"] = 22
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile("test.xlsx", buffer.read())
